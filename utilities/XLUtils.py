import openpyxl


class XLUtils:

    def getrowCount(file, sheetName):       # these are the reusable method which are created inside exel XLUtilize.py module
            workbook = openpyxl.load_workbook(file)
            sheet = workbook[sheetName]
            return sheet(sheet.max_row)   # it will return row count

    def getcolumnCount(file,sheetName):  # we can access these method by using exel utilities class itself and it will return required data.
            workbook = openpyxl.load_workbook(file)
            sheet = workbook[sheetName]
            return sheet(sheet.max_column)  # it will return column count

    def readData(file,sheetName,rownum,columnnum):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook[sheetName]
            return sheet.cell(row=rownum, column=columnnum).value  # it will read the data from exel

    def writeData(file,sheetName,rownum,columnnum,data):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook[sheetName]
            sheet.cell(row=rownum, column=columnnum).value = data  # it will write data into exel
            workbook.save(file)




